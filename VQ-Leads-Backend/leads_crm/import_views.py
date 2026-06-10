import csv
import io
import re
from decimal import Decimal, InvalidOperation

from django.contrib.auth.models import User
from django.core.validators import validate_email
from django.core.exceptions import ValidationError as DjangoValidationError
from django.http import HttpResponse
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Lead, LeadActivity, ImportHistory, ImportLog, ImportMappingTemplate, Notification
from .serializers import (
    ImportHistorySerializer, ImportHistoryDetailSerializer,
    ImportMappingTemplateSerializer
)

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - handled gracefully at runtime
    load_workbook = None


MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
PREVIEW_LIMIT = 20
BATCH_SIZE = 100

SYSTEM_FIELDS = ['name', 'phone', 'email', 'company', 'source', 'value', 'status']
LEAD_STATUS_VALUES = {key for key, _ in Lead.STATUS_CHOICES}


class IsAdminUserRole(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and hasattr(request.user, 'profile') and request.user.profile.role == 'ADMIN'


def normalize_header(value):
    if value is None:
        return ''
    return str(value).strip()


def normalize_phone(value):
    if value is None:
        return ''
    return re.sub(r'\D', '', str(value))


def normalize_email(value):
    if value is None:
        return ''
    return str(value).strip().lower()


def parse_uploaded_file(uploaded_file):
    filename = uploaded_file.name.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
        raise ValueError('Only CSV and XLSX files are supported.')

    if uploaded_file.size > MAX_FILE_SIZE:
        raise ValueError('File is too large. Maximum allowed size is 10MB.')

    if filename.endswith('.csv'):
        content = uploaded_file.read().decode('utf-8-sig', errors='ignore')
        reader = csv.DictReader(io.StringIO(content))
        headers = [normalize_header(h) for h in (reader.fieldnames or [])]
        rows = [{normalize_header(k): (v or '').strip() for k, v in row.items()} for row in reader]
        return headers, rows, 'csv'

    if load_workbook is None:
        raise ValueError('XLSX import requires openpyxl to be installed.')

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], [], 'xlsx'

    headers = [normalize_header(c) for c in header_row]
    rows = []
    for excel_row in rows_iter:
        row_dict = {}
        for idx, header in enumerate(headers):
            value = excel_row[idx] if idx < len(excel_row) else ''
            row_dict[header] = '' if value is None else str(value).strip()
        rows.append(row_dict)
    return headers, rows, 'xlsx'


def detect_mapping(headers):
    normalized = {h.lower(): h for h in headers}
    aliases = {
        'name': ['name', 'full name', 'customer name', 'lead name'],
        'phone': ['phone', 'mobile', 'mobile number', 'phone number', 'contact number'],
        'email': ['email', 'mail', 'mail id', 'email address'],
        'company': ['company', 'organization', 'business'],
        'source': ['source', 'lead source', 'channel'],
        'value': ['value', 'deal value', 'amount', 'lead value'],
        'status': ['status', 'lead status'],
    }
    mapping = {}
    for system_field, candidates in aliases.items():
        for key in candidates:
            if key in normalized:
                mapping[system_field] = normalized[key]
                break
    return mapping


def map_row(raw_row, mapping):
    return {field: (raw_row.get(column, '') or '').strip() for field, column in mapping.items()}


def is_empty_mapped_row(mapped_row):
    return all(not str(v).strip() for v in mapped_row.values())


def validate_mapped_row(mapped_row):
    errors = []
    name = (mapped_row.get('name') or '').strip()
    phone = (mapped_row.get('phone') or '').strip()
    email = (mapped_row.get('email') or '').strip()
    value = (mapped_row.get('value') or '').strip()

    if not name:
        errors.append('Name is required.')
    if not phone and not email:
        errors.append('Either phone or email is required.')
    if phone:
        digits = normalize_phone(phone)
        if len(digits) < 8 or len(digits) > 15:
            errors.append('Invalid phone number.')
    if email:
        try:
            validate_email(email)
        except DjangoValidationError:
            errors.append('Invalid email address.')
    if value:
        try:
            Decimal(str(value))
        except (InvalidOperation, ValueError):
            errors.append('Invalid lead value.')
    return errors


def lead_defaults_from_row(mapped_row):
    status_val = (mapped_row.get('status') or '').strip().upper()
    return {
        'name': mapped_row.get('name') or '',
        'phone': mapped_row.get('phone') or '',
        'email': mapped_row.get('email') or '',
        'company': mapped_row.get('company') or '',
        'source': mapped_row.get('source') or 'Imported',
        'value': Decimal(str(mapped_row.get('value') or '0') or '0'),
        'status': status_val if status_val in LEAD_STATUS_VALUES else 'NEW',
    }


def send_new_lead_available_notifications(lead):
    recipients = User.objects.filter(profile__role__in=['AGENT', 'LEADER'], is_active=True)
    Notification.objects.bulk_create([
        Notification(
            recipient=recipient,
            type='NEW_LEAD_AVAILABLE',
            title='New Lead Available',
            message=f"A new imported lead ({lead.name}) is available to claim.",
            lead=lead,
        )
        for recipient in recipients
    ])


class ImportPreviewView(APIView):
    permission_classes = [IsAdminUserRole]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'File is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            headers, rows, file_type = parse_uploaded_file(uploaded_file)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        mapping = detect_mapping(headers)
        preview_rows = []
        invalid_count = 0
        empty_count = 0
        for idx, raw in enumerate(rows[:PREVIEW_LIMIT], start=2):
            mapped = map_row(raw, mapping) if mapping else {}
            row_errors = validate_mapped_row(mapped) if mapped else ['Map required columns to validate rows.']
            is_empty = is_empty_mapped_row(mapped) if mapped else False
            if is_empty:
                empty_count += 1
            if row_errors and not is_empty:
                invalid_count += 1
            preview_rows.append({
                'rowNumber': idx,
                'raw': raw,
                'mapped': mapped,
                'isEmpty': is_empty,
                'errors': row_errors,
            })

        return Response({
            'fileName': uploaded_file.name,
            'fileSize': uploaded_file.size,
            'fileType': file_type,
            'headers': headers,
            'detectedMapping': mapping,
            'totalRecords': len(rows),
            'previewRows': preview_rows,
            'invalidRows': invalid_count,
            'emptyRows': empty_count,
        })


class ImportDuplicateCheckView(APIView):
    permission_classes = [IsAdminUserRole]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        uploaded_file = request.FILES.get('file')
        mapping_raw = request.data.get('mapping')
        if not uploaded_file:
            return Response({'error': 'File is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not mapping_raw:
            return Response({'error': 'Column mapping is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import json
            mapping = json.loads(mapping_raw)
        except Exception:
            return Response({'error': 'Invalid mapping payload.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            _, rows, _ = parse_uploaded_file(uploaded_file)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        existing_phones = {normalize_phone(l.phone): l.id for l in Lead.objects.exclude(phone='').only('id', 'phone')}
        existing_emails = {normalize_email(l.email): l.id for l in Lead.objects.exclude(email='').only('id', 'email')}

        duplicates = []
        for idx, raw in enumerate(rows, start=2):
            mapped = map_row(raw, mapping)
            phone_key = normalize_phone(mapped.get('phone'))
            email_key = normalize_email(mapped.get('email'))
            phone_match = existing_phones.get(phone_key) if phone_key else None
            email_match = existing_emails.get(email_key) if email_key else None
            if phone_match or email_match:
                duplicates.append({
                    'rowNumber': idx,
                    'name': mapped.get('name') or raw.get(mapping.get('name', ''), ''),
                    'phone': mapped.get('phone', ''),
                    'email': mapped.get('email', ''),
                    'matchedLeadId': phone_match or email_match,
                    'matchReason': 'phone' if phone_match else 'email',
                })

        return Response({
            'duplicateCount': len(duplicates),
            'duplicates': duplicates[:200],
        })


class ImportExecuteView(APIView):
    permission_classes = [IsAdminUserRole]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        uploaded_file = request.FILES.get('file')
        mapping_raw = request.data.get('mapping')
        duplicate_strategy = (request.data.get('duplicateStrategy') or 'SKIP').upper()
        if duplicate_strategy not in {'SKIP', 'UPDATE', 'IMPORT_ALL'}:
            return Response({'error': 'Invalid duplicate strategy.'}, status=status.HTTP_400_BAD_REQUEST)
        if not uploaded_file or not mapping_raw:
            return Response({'error': 'File and mapping are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import json
            mapping = json.loads(mapping_raw)
        except Exception:
            return Response({'error': 'Invalid mapping payload.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            _, rows, file_type = parse_uploaded_file(uploaded_file)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        history = ImportHistory.objects.create(
            file_name=uploaded_file.name,
            file_type=file_type,
            total_records=len(rows),
            imported_by=request.user,
            status='PROCESSING',
            duplicate_strategy=duplicate_strategy,
            column_mapping=mapping,
        )

        existing_by_phone = {normalize_phone(l.phone): l for l in Lead.objects.exclude(phone='')}
        existing_by_email = {normalize_email(l.email): l for l in Lead.objects.exclude(email='')}

        success_count = failed_count = duplicate_count = 0
        logs_to_create = []

        for idx, raw in enumerate(rows, start=2):
            mapped = map_row(raw, mapping)
            if is_empty_mapped_row(mapped):
                failed_count += 1
                logs_to_create.append(ImportLog(
                    import_history=history,
                    row_number=idx,
                    status='FAILED',
                    error_message='Empty row.',
                    row_data=mapped,
                ))
                continue

            errors = validate_mapped_row(mapped)
            if errors:
                failed_count += 1
                logs_to_create.append(ImportLog(
                    import_history=history,
                    row_number=idx,
                    status='FAILED',
                    error_message='; '.join(errors),
                    row_data=mapped,
                ))
                continue

            phone_key = normalize_phone(mapped.get('phone'))
            email_key = normalize_email(mapped.get('email'))
            existing_lead = (existing_by_phone.get(phone_key) if phone_key else None) or (existing_by_email.get(email_key) if email_key else None)

            if existing_lead and duplicate_strategy == 'SKIP':
                duplicate_count += 1
                logs_to_create.append(ImportLog(
                    import_history=history,
                    row_number=idx,
                    status='DUPLICATE',
                    error_message='Duplicate lead found. Skipped.',
                    row_data=mapped,
                ))
            elif existing_lead and duplicate_strategy == 'UPDATE':
                duplicate_count += 1
                data = lead_defaults_from_row(mapped)
                for key, value in data.items():
                    setattr(existing_lead, key, value)
                existing_lead.save()
                success_count += 1
                logs_to_create.append(ImportLog(
                    import_history=history,
                    row_number=idx,
                    status='UPDATED',
                    error_message='',
                    row_data=mapped,
                ))
            else:
                try:
                    data = lead_defaults_from_row(mapped)
                    lead = Lead.objects.create(**data, owner=None)
                    send_new_lead_available_notifications(lead)
                    LeadActivity.objects.create(
                        lead=lead,
                        user=request.user,
                        activity_type='CREATED',
                        description=f"Lead imported via bulk import by {request.user.username}.",
                    )
                    success_count += 1
                    logs_to_create.append(ImportLog(
                        import_history=history,
                        row_number=idx,
                        status='SUCCESS',
                        error_message='',
                        row_data=mapped,
                    ))
                except Exception as exc:  # pragma: no cover
                    failed_count += 1
                    logs_to_create.append(ImportLog(
                        import_history=history,
                        row_number=idx,
                        status='FAILED',
                        error_message=f'Import failed: {exc}',
                        row_data=mapped,
                    ))

            if len(logs_to_create) >= BATCH_SIZE:
                ImportLog.objects.bulk_create(logs_to_create)
                logs_to_create = []
                ImportHistory.objects.filter(pk=history.pk).update(
                    success_count=success_count,
                    failed_count=failed_count,
                    duplicate_count=duplicate_count,
                )

        if logs_to_create:
            ImportLog.objects.bulk_create(logs_to_create)

        final_status = 'COMPLETED'
        if failed_count > 0 and success_count > 0:
            final_status = 'PARTIAL'
        elif failed_count > 0 and success_count == 0:
            final_status = 'FAILED'

        history.success_count = success_count
        history.failed_count = failed_count
        history.duplicate_count = duplicate_count
        history.status = final_status
        history.save()

        return Response(ImportHistoryDetailSerializer(history).data, status=status.HTTP_201_CREATED)


class ImportHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAdminUserRole]
    queryset = ImportHistory.objects.all().select_related('imported_by').order_by('-created_at')

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ImportHistoryDetailSerializer
        return ImportHistorySerializer

    def get_queryset(self):
        qs = super().get_queryset()
        status_filter = self.request.query_params.get('status')
        search = self.request.query_params.get('search')
        if status_filter:
            qs = qs.filter(status=status_filter)
        if search:
            qs = qs.filter(file_name__icontains=search)
        return qs

    @action(detail=True, methods=['get'], url_path='error-report')
    def error_report(self, request, pk=None):
        history = self.get_object()
        logs = history.logs.filter(status__in=['FAILED', 'DUPLICATE']).order_by('row_number')

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['row_number', 'status', 'error_message', 'row_data'])
        for log in logs:
            writer.writerow([log.row_number, log.status, log.error_message, log.row_data])

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=import_errors_{history.id}.csv'
        return response

    @action(detail=True, methods=['post'], url_path='retry-failed')
    def retry_failed(self, request, pk=None):
        history = self.get_object()
        failed_logs = history.logs.filter(status='FAILED').order_by('row_number')
        if not failed_logs.exists():
            return Response({'error': 'No failed rows to retry.'}, status=status.HTTP_400_BAD_REQUEST)

        retry = ImportHistory.objects.create(
            file_name=f"{history.file_name} (retry)",
            file_type=history.file_type,
            total_records=failed_logs.count(),
            imported_by=request.user,
            status='PROCESSING',
            duplicate_strategy=history.duplicate_strategy,
            column_mapping=history.column_mapping,
        )

        existing_by_phone = {normalize_phone(l.phone): l for l in Lead.objects.exclude(phone='')}
        existing_by_email = {normalize_email(l.email): l for l in Lead.objects.exclude(email='')}

        success_count = failed_count = duplicate_count = 0
        logs_to_create = []

        for failed in failed_logs:
            mapped = failed.row_data or {}
            errors = validate_mapped_row(mapped)
            if errors:
                failed_count += 1
                logs_to_create.append(ImportLog(
                    import_history=retry,
                    row_number=failed.row_number,
                    status='FAILED',
                    error_message='; '.join(errors),
                    row_data=mapped,
                ))
                continue

            phone_key = normalize_phone(mapped.get('phone'))
            email_key = normalize_email(mapped.get('email'))
            existing_lead = (existing_by_phone.get(phone_key) if phone_key else None) or (existing_by_email.get(email_key) if email_key else None)

            if existing_lead and retry.duplicate_strategy == 'SKIP':
                duplicate_count += 1
                logs_to_create.append(ImportLog(
                    import_history=retry,
                    row_number=failed.row_number,
                    status='DUPLICATE',
                    error_message='Duplicate lead found. Skipped.',
                    row_data=mapped,
                ))
                continue

            if existing_lead and retry.duplicate_strategy == 'UPDATE':
                duplicate_count += 1
                data = lead_defaults_from_row(mapped)
                for key, value in data.items():
                    setattr(existing_lead, key, value)
                existing_lead.save()
                success_count += 1
                logs_to_create.append(ImportLog(
                    import_history=retry,
                    row_number=failed.row_number,
                    status='UPDATED',
                    error_message='',
                    row_data=mapped,
                ))
                continue

            try:
                data = lead_defaults_from_row(mapped)
                lead = Lead.objects.create(**data, owner=None)
                send_new_lead_available_notifications(lead)
                LeadActivity.objects.create(
                    lead=lead,
                    user=request.user,
                    activity_type='CREATED',
                    description=f"Lead imported via retry by {request.user.username}.",
                )
                success_count += 1
                logs_to_create.append(ImportLog(
                    import_history=retry,
                    row_number=failed.row_number,
                    status='SUCCESS',
                    error_message='',
                    row_data=mapped,
                ))
            except Exception as exc:  # pragma: no cover
                failed_count += 1
                logs_to_create.append(ImportLog(
                    import_history=retry,
                    row_number=failed.row_number,
                    status='FAILED',
                    error_message=f'Retry failed: {exc}',
                    row_data=mapped,
                ))

        if logs_to_create:
            ImportLog.objects.bulk_create(logs_to_create)

        retry.success_count = success_count
        retry.failed_count = failed_count
        retry.duplicate_count = duplicate_count
        if failed_count and success_count:
            retry.status = 'PARTIAL'
        elif failed_count and not success_count:
            retry.status = 'FAILED'
        else:
            retry.status = 'COMPLETED'
        retry.save()

        return Response(ImportHistoryDetailSerializer(retry).data)


class ImportMappingTemplateViewSet(viewsets.ModelViewSet):
    serializer_class = ImportMappingTemplateSerializer
    permission_classes = [IsAdminUserRole]

    def get_queryset(self):
        return ImportMappingTemplate.objects.filter(created_by=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
