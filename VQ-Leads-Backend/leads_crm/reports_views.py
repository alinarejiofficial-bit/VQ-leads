import csv
import io
from datetime import timedelta
from decimal import Decimal

from django.contrib.auth.models import User
from django.db.models import Avg, Count, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import (
    Commission, CommissionReport, ExportHistory, FollowUp, Lead,
    LeadActivity, ReportCache, CallLog
)
from .serializers import ExportHistorySerializer

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover
    Workbook = None

try:
    from fpdf import FPDF
except Exception:  # pragma: no cover
    FPDF = None


class IsLeaderOrAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return (
            request.user.is_authenticated
            and hasattr(request.user, 'profile')
            and request.user.profile.role in ('ADMIN', 'LEADER')
        )


def parse_filters(request):
    quick = request.query_params.get('quickFilter') or 'THIS_MONTH'
    custom_from = request.query_params.get('dateFrom')
    custom_to = request.query_params.get('dateTo')

    now = timezone.now()
    start = end = None
    if quick == 'TODAY':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif quick == 'YESTERDAY':
        y = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        start = y
        end = y + timedelta(days=1)
    elif quick == 'THIS_WEEK':
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif quick == 'THIS_MONTH':
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif quick == 'THIS_QUARTER':
        q_month = ((now.month - 1) // 3) * 3 + 1
        start = now.replace(month=q_month, day=1, hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif quick == 'THIS_YEAR':
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif quick == 'CUSTOM' and custom_from and custom_to:
        try:
            from datetime import datetime
            start = timezone.make_aware(datetime.fromisoformat(custom_from))
            end = timezone.make_aware(datetime.fromisoformat(custom_to))
        except Exception:
            start = end = None

    return {
        'start': start,
        'end': end,
        'lead_status': request.query_params.get('leadStatus') or '',
        'lead_source': request.query_params.get('leadSource') or '',
        'agent': request.query_params.get('agent') or '',
        'team_member': request.query_params.get('teamMember') or '',
        'campaign': request.query_params.get('campaign') or '',
        'region': request.query_params.get('region') or '',
        'quick_filter': quick,
    }


def apply_lead_filters(qs, filters):
    if filters['start'] and filters['end']:
        qs = qs.filter(created_at__range=[filters['start'], filters['end']])
    if filters['lead_status']:
        qs = qs.filter(status=filters['lead_status'])
    if filters['lead_source']:
        qs = qs.filter(source=filters['lead_source'])
    if filters['agent']:
        qs = qs.filter(owner_id=filters['agent'])
    if filters['team_member']:
        qs = qs.filter(owner_id=filters['team_member'])
    if filters['campaign']:
        qs = qs.filter(source__icontains=filters['campaign'])
    if filters['region']:
        qs = qs.filter(company__icontains=filters['region'])
    return qs


def apply_date_window(qs, date_field, filters):
    if filters['start'] and filters['end']:
        return qs.filter(**{f'{date_field}__range': [filters['start'], filters['end']]})
    return qs


def as_float(v):
    if v is None:
        return 0.0
    return float(v)


def lead_report_data(filters):
    qs = apply_lead_filters(Lead.objects.select_related('owner').all(), filters)
    total = qs.count()
    status_counts = {item['status']: item['count'] for item in qs.values('status').annotate(count=Count('id'))}

    open_qs = qs.exclude(status__in=['WON', 'LOST'])
    aging_rows = []
    for lead in open_qs:
        age_days = (timezone.now() - lead.created_at).days
        aging_rows.append({
            'leadName': lead.name,
            'status': lead.status,
            'owner': lead.owner.get_full_name() if lead.owner else 'Unassigned',
            'source': lead.source,
            'ageDays': age_days,
            'createdAt': lead.created_at.isoformat(),
        })

    daily_growth = [
        {'date': d['day'].strftime('%b %d'), 'count': d['count']}
        for d in apply_date_window(qs, 'created_at', filters)
        .annotate(day=TruncDate('created_at')).values('day').annotate(count=Count('id')).order_by('day')[:31]
    ]

    pipeline_order = ['NEW', 'CONTACTED', 'IN_PROGRESS', 'QUALIFIED', 'PROPOSAL_SENT', 'NEGOTIATION', 'WON']
    pipeline = [{'stage': s, 'count': status_counts.get(s, 0)} for s in pipeline_order]

    return {
        'metrics': {
            'totalLeads': total,
            'newLeads': status_counts.get('NEW', 0),
            'contactedLeads': status_counts.get('CONTACTED', 0),
            'qualifiedLeads': status_counts.get('QUALIFIED', 0),
            'proposalSentLeads': status_counts.get('PROPOSAL_SENT', 0),
            'negotiationLeads': status_counts.get('NEGOTIATION', 0),
            'wonLeads': status_counts.get('WON', 0),
            'lostLeads': status_counts.get('LOST', 0),
        },
        'statusBreakdown': status_counts,
        'growthTrend': daily_growth,
        'pipelineFunnel': pipeline,
        'leadAging': sorted(aging_rows, key=lambda x: x['ageDays'], reverse=True)[:200],
        'pipelineValue': as_float(qs.aggregate(v=Sum('value'))['v']),
    }


def conversion_report_data(filters):
    qs = apply_lead_filters(Lead.objects.select_related('owner').all(), filters)
    total = qs.count()
    won_qs = qs.filter(status='WON')
    lost_qs = qs.filter(status='LOST')

    conversion_rate = round((won_qs.count() / total) * 100, 2) if total else 0.0

    avg_conversion_days = 0.0
    if won_qs.exists():
        total_days = sum((lead.updated_at - lead.created_at).days for lead in won_qs)
        avg_conversion_days = round(total_days / won_qs.count(), 1)

    monthly = [
        {'month': m['month'].strftime('%b %Y'), 'converted': m['count']}
        for m in apply_date_window(won_qs, 'updated_at', filters)
        .annotate(month=TruncMonth('updated_at')).values('month').annotate(count=Count('id')).order_by('month')[:12]
    ]

    by_agent = []
    for row in won_qs.values('owner__username', 'owner__first_name', 'owner__last_name').annotate(count=Count('id')).order_by('-count')[:20]:
        name = f"{row['owner__first_name']} {row['owner__last_name']}".strip() or (row['owner__username'] or 'Unassigned')
        by_agent.append({'agent': name, 'username': row['owner__username'] or '', 'converted': row['count']})

    by_source = [
        {'source': r['source'], 'converted': r['count']}
        for r in won_qs.values('source').annotate(count=Count('id')).order_by('-count')[:12]
    ]

    return {
        'metrics': {
            'totalLeads': total,
            'convertedLeads': won_qs.count(),
            'conversionRate': conversion_rate,
            'lostLeads': lost_qs.count(),
            'avgConversionTimeDays': avg_conversion_days,
        },
        'funnel': [
            {'stage': 'NEW', 'count': qs.filter(status='NEW').count()},
            {'stage': 'CONTACTED', 'count': qs.filter(status='CONTACTED').count()},
            {'stage': 'QUALIFIED', 'count': qs.filter(status='QUALIFIED').count()},
            {'stage': 'PROPOSAL_SENT', 'count': qs.filter(status='PROPOSAL_SENT').count()},
            {'stage': 'NEGOTIATION', 'count': qs.filter(status='NEGOTIATION').count()},
            {'stage': 'WON', 'count': won_qs.count()},
        ],
        'monthlyTrend': monthly,
        'conversionByAgent': by_agent,
        'conversionBySource': by_source,
    }


def source_report_data(filters):
    qs = apply_lead_filters(Lead.objects.all(), filters)
    rows = []
    for row in qs.values('source').annotate(
        leads=Count('id'),
        converted=Count('id', filter=Q(status='WON')),
        revenue=Sum('value', filter=Q(status='WON')),
    ).order_by('-leads')[:20]:
        leads = row['leads'] or 0
        converted = row['converted'] or 0
        rate = round((converted / leads) * 100, 2) if leads else 0.0
        rows.append({
            'source': row['source'] or 'Unknown',
            'leadsGenerated': leads,
            'convertedLeads': converted,
            'conversionRate': rate,
            'revenueGenerated': as_float(row['revenue']),
        })

    return {
        'table': rows,
        'sourceComparison': [{'label': r['source'], 'value': r['leadsGenerated']} for r in rows],
        'sourceDistribution': [{'label': r['source'], 'value': r['leadsGenerated']} for r in rows],
        'revenueBySource': [{'label': r['source'], 'value': r['revenueGenerated']} for r in rows],
    }


def team_report_data(filters):
    users_qs = User.objects.filter(profile__role__in=['AGENT', 'LEADER'], is_active=True)
    if filters['agent']:
        users_qs = users_qs.filter(id=filters['agent'])

    users = list(users_qs)
    rows = []
    for user in users:
        assigned_qs = apply_lead_filters(Lead.objects.filter(owner=user), filters)
        followups_qs = apply_date_window(FollowUp.objects.filter(assigned_agent=user), 'scheduled_time', filters)
        calls_qs = apply_date_window(CallLog.objects.filter(agent=user), 'call_date', filters)
        meetings_done = followups_qs.filter(status='COMPLETED', followup_type='MEETING').count()
        won_qs = assigned_qs.filter(status='WON')
        assigned = assigned_qs.count()
        won = won_qs.count()
        conversion_rate = round((won / assigned) * 100, 2) if assigned else 0.0
        revenue = as_float(won_qs.aggregate(v=Sum('value'))['v'])
        activities_completed = (
            followups_qs.filter(status='COMPLETED').count()
            + calls_qs.count()
            + apply_date_window(LeadActivity.objects.filter(user=user), 'created_at', filters).count()
        )
        rows.append({
            'agentName': user.get_full_name() or user.username,
            'username': user.username,
            'assignedLeads': assigned,
            'leadsContacted': assigned_qs.filter(status__in=['CONTACTED', 'IN_PROGRESS', 'QUALIFIED', 'WON']).count(),
            'followupsCompleted': followups_qs.filter(status='COMPLETED').count(),
            'callsMade': calls_qs.count(),
            'meetingsConducted': meetings_done,
            'wonDeals': won,
            'revenueGenerated': revenue,
            'activitiesCompleted': activities_completed,
            'conversionRate': conversion_rate,
        })

    rows.sort(key=lambda x: x['revenueGenerated'], reverse=True)

    return {
        'table': rows,
        'leaderboard': [{'label': r['agentName'], 'value': r['revenueGenerated']} for r in rows[:10]],
        'agentComparison': [{'label': r['agentName'], 'value': r['conversionRate']} for r in rows[:10]],
        'revenueByAgent': [{'label': r['agentName'], 'value': r['revenueGenerated']} for r in rows[:10]],
    }


def commission_report_data(filters):
    qs = Commission.objects.select_related('agent', 'lead')
    if filters['start'] and filters['end']:
        qs = qs.filter(calculated_at__range=[filters['start'], filters['end']])
    if filters['agent']:
        qs = qs.filter(agent_id=filters['agent'])

    total_sales = as_float(qs.aggregate(v=Sum('lead__value'))['v'])
    commission_earned = as_float(qs.aggregate(v=Sum('amount'))['v'])
    paid_commission = as_float(qs.filter(status='PAID').aggregate(v=Sum('amount'))['v'])
    pending_commission = as_float(qs.filter(status='PENDING').aggregate(v=Sum('amount'))['v'])

    # mirror into commission_reports table for reporting cache snapshots
    today = timezone.now().date()
    for row in qs.values('agent', 'agent__username').annotate(
        sales=Sum('lead__value'),
        amount=Sum('amount'),
        avg_rate=Avg('rate'),
    ):
        if not CommissionReport.objects.filter(agent_id=row['agent'], created_at__date=today).exists():
            CommissionReport.objects.create(
                agent_id=row['agent'],
                sales_amount=row['sales'] or Decimal('0.00'),
                commission_rate=row['avg_rate'] or Decimal('0.00'),
                commission_amount=row['amount'] or Decimal('0.00'),
                payment_status='PAID' if paid_commission > 0 and pending_commission == 0 else 'PENDING',
            )

    monthly = [
        {
            'month': m['month'].strftime('%b %Y'),
            'commissionEarned': as_float(m['earned']),
            'commissionPaid': as_float(m['paid']),
        }
        for m in qs.annotate(month=TruncMonth('calculated_at')).values('month').annotate(
            earned=Sum('amount'),
            paid=Sum('amount', filter=Q(status='PAID')),
        ).order_by('month')[:12]
    ]

    agent_rows = [
        {
            'agent': (f"{a['agent__first_name']} {a['agent__last_name']}".strip() or a['agent__username']),
            'username': a['agent__username'],
            'salesAmount': as_float(a['sales']),
            'commissionRate': as_float(a['avg_rate']),
            'commissionAmount': as_float(a['earned']),
            'paymentStatus': 'PENDING' if a['pending_count'] else 'PAID',
        }
        for a in qs.values('agent__username', 'agent__first_name', 'agent__last_name').annotate(
            sales=Sum('lead__value'),
            earned=Sum('amount'),
            avg_rate=Avg('rate'),
            pending_count=Count('id', filter=Q(status='PENDING')),
        ).order_by('-earned')
    ]

    return {
        'metrics': {
            'totalSales': total_sales,
            'commissionEarned': commission_earned,
            'paidCommission': paid_commission,
            'pendingCommission': pending_commission,
        },
        'monthlyReport': monthly,
        'agentWise': agent_rows,
    }


def dashboard_widgets(filters):
    leads_qs = apply_lead_filters(Lead.objects.all(), filters)
    total_leads = leads_qs.count()
    won_qs = leads_qs.filter(status='WON')
    lost_qs = leads_qs.filter(status='LOST')
    conversion_rate = round((won_qs.count() / total_leads) * 100, 2) if total_leads else 0.0

    source_row = leads_qs.values('source').annotate(c=Count('id')).order_by('-c').first()
    top_source = source_row['source'] if source_row else 'N/A'

    top_agent_row = leads_qs.filter(owner__isnull=False).values('owner__username', 'owner__first_name', 'owner__last_name').annotate(
        won=Count('id', filter=Q(status='WON'))
    ).order_by('-won').first()
    if top_agent_row:
        top_agent = (f"{top_agent_row['owner__first_name']} {top_agent_row['owner__last_name']}".strip() or top_agent_row['owner__username'])
    else:
        top_agent = 'N/A'

    now = timezone.now()
    followups_qs = FollowUp.objects.filter(status='SCHEDULED')
    if filters['agent']:
        followups_qs = followups_qs.filter(assigned_agent_id=filters['agent'])
    pending_followups = followups_qs.filter(scheduled_time__gte=now).count()

    comm_qs = Commission.objects.all()
    if filters['start'] and filters['end']:
        comm_qs = comm_qs.filter(calculated_at__range=[filters['start'], filters['end']])
    if filters['agent']:
        comm_qs = comm_qs.filter(agent_id=filters['agent'])

    return {
        'totalLeads': total_leads,
        'pipelineValue': as_float(leads_qs.aggregate(v=Sum('value'))['v']),
        'conversionRate': conversion_rate,
        'revenueGenerated': as_float(won_qs.aggregate(v=Sum('value'))['v']),
        'topLeadSource': top_source,
        'topPerformingAgent': top_agent,
        'pendingFollowups': pending_followups,
        'wonDeals': won_qs.count(),
        'lostDeals': lost_qs.count(),
        'commissionPaid': as_float(comm_qs.filter(status='PAID').aggregate(v=Sum('amount'))['v']),
    }


def flatten_rows(data):
    if isinstance(data, dict):
        if 'table' in data and isinstance(data['table'], list):
            return data['table']
        if 'agentWise' in data and isinstance(data['agentWise'], list):
            return data['agentWise']
        if 'monthlyReport' in data and isinstance(data['monthlyReport'], list):
            return data['monthlyReport']
    if isinstance(data, list):
        return data
    return []


def to_csv_bytes(rows):
    if not rows:
        output = io.StringIO()
        output.write('No data\n')
        return output.getvalue().encode('utf-8')
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode('utf-8')


def to_xlsx_bytes(rows):
    if Workbook is None:
        raise ValueError('XLSX export requires openpyxl.')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, '') for h in headers])
    else:
        ws.append(['No data'])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def to_pdf_bytes(rows, title):
    if FPDF is None:
        raise ValueError('PDF export requires fpdf2.')
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, f'{title} Report', ln=True)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 7, f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}', ln=True)
    pdf.ln(2)

    if not rows:
        pdf.cell(0, 7, 'No data available', ln=True)
        return bytes(pdf.output(dest='S'))

    headers = list(rows[0].keys())
    col_width = 270 / max(len(headers), 1)
    pdf.set_font('Helvetica', 'B', 8)
    for h in headers:
        pdf.cell(col_width, 7, str(h)[:22], border=1)
    pdf.ln()
    pdf.set_font('Helvetica', '', 7)
    for row in rows:
        for h in headers:
            pdf.cell(col_width, 6, str(row.get(h, ''))[:22], border=1)
        pdf.ln()
    return bytes(pdf.output(dest='S'))


class ReportsDashboardWidgetsView(APIView):
    permission_classes = [IsLeaderOrAdmin]

    def get(self, request):
        filters = parse_filters(request)
        data = dashboard_widgets(filters)
        ReportCache.objects.create(report_type='dashboard_widgets', report_data=data, generated_by=request.user)
        return Response(data)


class ReportsAnalyticsView(APIView):
    permission_classes = [IsLeaderOrAdmin]

    def get(self, request):
        report_type = (request.query_params.get('reportType') or 'lead').lower()
        filters = parse_filters(request)

        if report_type == 'lead':
            data = lead_report_data(filters)
        elif report_type == 'conversion':
            data = conversion_report_data(filters)
        elif report_type == 'source':
            data = source_report_data(filters)
        elif report_type == 'team':
            data = team_report_data(filters)
        elif report_type == 'commission':
            data = commission_report_data(filters)
        else:
            return Response({'error': 'Unsupported report type.'}, status=400)

        ReportCache.objects.create(report_type=report_type, report_data=data, generated_by=request.user)
        return Response(data)


class ReportsExportView(APIView):
    permission_classes = [IsLeaderOrAdmin]

    def post(self, request):
        report_type = (request.data.get('reportType') or 'lead').lower()
        file_type = (request.data.get('fileType') or 'csv').lower()
        if file_type not in {'csv', 'xlsx', 'pdf'}:
            return Response({'error': 'Unsupported format.'}, status=400)

        # Reuse GET pipeline by mapping body filters to query-style parsing
        class _Obj:
            query_params = {
                'quickFilter': request.data.get('quickFilter', 'THIS_MONTH'),
                'dateFrom': request.data.get('dateFrom', ''),
                'dateTo': request.data.get('dateTo', ''),
                'leadStatus': request.data.get('leadStatus', ''),
                'leadSource': request.data.get('leadSource', ''),
                'agent': request.data.get('agent', ''),
                'teamMember': request.data.get('teamMember', ''),
                'campaign': request.data.get('campaign', ''),
                'region': request.data.get('region', ''),
            }

        filters = parse_filters(_Obj())

        if report_type == 'lead':
            data = lead_report_data(filters)
        elif report_type == 'conversion':
            data = conversion_report_data(filters)
        elif report_type == 'source':
            data = source_report_data(filters)
        elif report_type == 'team':
            data = team_report_data(filters)
        elif report_type == 'commission':
            data = commission_report_data(filters)
        else:
            return Response({'error': 'Unsupported report type.'}, status=400)

        rows = flatten_rows(data)

        if file_type == 'csv':
            content = to_csv_bytes(rows)
            content_type = 'text/csv'
        elif file_type == 'xlsx':
            content = to_xlsx_bytes(rows)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            content = to_pdf_bytes(rows, report_type.title())
            content_type = 'application/pdf'

        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        file_name = f'{report_type}_report_{timestamp}.{file_type}'
        history = ExportHistory.objects.create(
            file_name=file_name,
            file_type=file_type,
            exported_by=request.user,
            total_records=len(rows),
            filters_applied={'reportType': report_type, 'filters': request.data},
            status='COMPLETED',
        )

        response = HttpResponse(content, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        response['X-Export-History-Id'] = str(history.id)
        return response


class ReportsExportHistoryView(APIView):
    permission_classes = [IsLeaderOrAdmin]

    def get(self, request):
        qs = ExportHistory.objects.order_by('-created_at')
        report_only = request.query_params.get('reportOnly', 'true').lower() == 'true'
        if report_only:
            qs = qs.filter(filters_applied__has_key='reportType')
        search = request.query_params.get('search')
        if search:
            qs = qs.filter(file_name__icontains=search)
        serializer = ExportHistorySerializer(qs[:100], many=True)
        return Response(serializer.data)
