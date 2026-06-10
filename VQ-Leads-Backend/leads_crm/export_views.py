import csv
import io
from datetime import timedelta
from decimal import Decimal

from django.contrib.auth.models import User
from django.db.models import Count, Q
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Lead, FollowUp, ExportHistory
from .serializers import ExportHistorySerializer
from .audit import log_audit

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover
    Workbook = None

try:
    from fpdf import FPDF
except Exception:  # pragma: no cover
    FPDF = None


EXPORT_COLUMNS = [
    'Lead Name',
    'Phone Number',
    'Email Address',
    'Company Name',
    'Lead Source',
    'Lead Status',
    'Assigned Agent',
    'Follow-Up Date',
    'Created Date',
    'Updated Date',
    'Expected Deal Value',
]

class IsAdminUserRole(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and hasattr(request.user, 'profile') and request.user.profile.role == 'ADMIN'


def parse_date_range(filters):
    date_range = (filters or {}).get('dateRange', 'ALL')
    now = timezone.now()
    if date_range == 'TODAY':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return start, now
    if date_range == 'THIS_WEEK':
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        return start, now
    if date_range == 'THIS_MONTH':
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, now
    if date_range == 'CUSTOM':
        custom_start = (filters or {}).get('customStartDate')
        custom_end = (filters or {}).get('customEndDate')
        if custom_start and custom_end:
            try:
                from datetime import datetime
                start = timezone.make_aware(datetime.fromisoformat(custom_start))
                end = timezone.make_aware(datetime.fromisoformat(custom_end))
                return start, end
            except Exception:
                return None, None
    return None, None


def apply_export_filters(qs, filters):
    filters = filters or {}

    status_values = [s for s in (filters.get('statuses') or []) if s]
    if status_values:
        qs = qs.filter(status__in=status_values)

    sources = [s for s in (filters.get('sources') or []) if s]
    if sources:
        qs = qs.filter(source__in=sources)

    team_member = filters.get('teamMember')
    if team_member:
        qs = qs.filter(owner_id=team_member)

    start, end = parse_date_range(filters)
    if start and end:
        qs = qs.filter(created_at__range=[start, end])

    return qs


def apply_export_scope(qs, payload):
    mode = (payload.get('exportMode') or 'FILTERED').upper()
    selected_ids = payload.get('selectedIds') or []
    current_page_ids = payload.get('currentPageIds') or []

    if mode == 'SELECTED':
        return qs.filter(id__in=selected_ids)
    if mode == 'CURRENT_PAGE':
        return qs.filter(id__in=current_page_ids)
    if mode in {'ALL', 'COMPLETE_DATASET', 'FILTERED'}:
        return qs
    return qs


def build_export_queryset(payload):
    mode = (payload.get('exportMode') or 'FILTERED').upper()
    qs = Lead.objects.select_related('owner').all()
    if mode not in {'ALL', 'COMPLETE_DATASET'}:
        qs = apply_export_filters(qs, payload.get('filters') or {})
    return apply_export_scope(qs, payload)


def nearest_followup_for_leads(lead_ids):
    followups = FollowUp.objects.filter(
        lead_id__in=lead_ids,
        completed=False
    ).order_by('lead_id', 'scheduled_time')
    result = {}
    for f in followups:
        if f.lead_id not in result:
            result[f.lead_id] = f.scheduled_time
    return result


def rows_from_leads(leads):
    lead_ids = [l.id for l in leads]
    followup_map = nearest_followup_for_leads(lead_ids)
    rows = []
    for lead in leads:
        owner_name = lead.owner.get_full_name().strip() if lead.owner else 'Unassigned'
        if lead.owner and not owner_name:
            owner_name = lead.owner.username
        followup = followup_map.get(lead.id)
        rows.append([
            lead.name,
            lead.phone,
            lead.email,
            lead.company,
            lead.source,
            lead.status,
            owner_name,
            followup.strftime('%Y-%m-%d %H:%M') if followup else '',
            lead.created_at.strftime('%Y-%m-%d %H:%M'),
            lead.updated_at.strftime('%Y-%m-%d %H:%M'),
            f"{Decimal(str(lead.value or '0')):.2f}",
        ])
    return rows


def generate_csv_bytes(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(EXPORT_COLUMNS)
    writer.writerows(rows)
    return output.getvalue().encode('utf-8')


def generate_xlsx_bytes(rows):
    if Workbook is None:
        raise ValueError('XLSX export requires openpyxl.')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leads Export'
    ws.append(EXPORT_COLUMNS)
    for row in rows:
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf_bytes(rows, filters):
    if FPDF is None:
        raise ValueError('PDF export requires fpdf2.')

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=10)

    # Company logo/title area
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'VQ Leads CRM', ln=True)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 8, 'Leads Export Report', ln=True)
    pdf.cell(0, 8, f"Export Date: {timezone.now().strftime('%Y-%m-%d %H:%M')}", ln=True)
    pdf.multi_cell(0, 7, f"Applied Filters: {filters or {}}")
    pdf.cell(0, 7, f"Total Records: {len(rows)}", ln=True)
    pdf.ln(2)

    # Summary statistics
    won = sum(1 for r in rows if r[5] == 'WON')
    lost = sum(1 for r in rows if r[5] == 'LOST')
    pdf.cell(0, 7, f"Summary: Won={won}, Lost={lost}, Open={len(rows) - won - lost}", ln=True)
    pdf.ln(2)

    # Table headers
    pdf.set_font('Helvetica', 'B', 8)
    col_widths = [30, 22, 35, 28, 24, 22, 28, 26, 24, 24, 24]
    for i, h in enumerate(EXPORT_COLUMNS):
        pdf.cell(col_widths[i], 7, h[:20], border=1)
    pdf.ln()

    pdf.set_font('Helvetica', '', 7)
    for row in rows:
        for i, cell in enumerate(row):
            pdf.cell(col_widths[i], 6, str(cell)[:20], border=1)
        pdf.ln()

    return bytes(pdf.output(dest='S'))


def file_response(file_name, file_type, rows, filters):
    if file_type == 'csv':
        data = generate_csv_bytes(rows)
        content_type = 'text/csv'
    elif file_type == 'xlsx':
        data = generate_xlsx_bytes(rows)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif file_type == 'pdf':
        data = generate_pdf_bytes(rows, filters)
        content_type = 'application/pdf'
    else:
        raise ValueError('Unsupported export format')

    response = HttpResponse(data, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response


class ExportPreviewView(APIView):
    permission_classes = [IsAdminUserRole]

    def post(self, request):
        qs = build_export_queryset(request.data)
        leads = list(qs.order_by('-created_at')[:5000])

        status_counts_qs = qs.values('status').annotate(count=Count('id'))
        status_counts = {item['status']: item['count'] for item in status_counts_qs}

        return Response({
            'totalRecords': len(leads),
            'sampleRows': rows_from_leads(leads[:20]),
            'summary': {
                'won': status_counts.get('WON', 0),
                'lost': status_counts.get('LOST', 0),
                'open': len(leads) - status_counts.get('WON', 0) - status_counts.get('LOST', 0),
            },
        })


class ExportGenerateView(APIView):
    permission_classes = [IsAdminUserRole]

    def post(self, request):
        filters = request.data.get('filters') or {}
        file_type = (request.data.get('fileType') or 'csv').lower()
        if file_type not in {'csv', 'xlsx', 'pdf'}:
            return Response({'error': 'Unsupported format'}, status=status.HTTP_400_BAD_REQUEST)

        qs = build_export_queryset(request.data)
        total = qs.count()
        if total == 0:
            return Response({'error': 'No leads match the selected export criteria.'}, status=status.HTTP_400_BAD_REQUEST)
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        file_name = f"leads_export_{timestamp}.{file_type}"

        history = ExportHistory.objects.create(
            file_name=file_name,
            file_type=file_type,
            exported_by=request.user,
            total_records=total,
            filters_applied={
                'filters': filters,
                'exportMode': request.data.get('exportMode') or 'FILTERED',
                'selectedIds': request.data.get('selectedIds') or [],
                'currentPageIds': request.data.get('currentPageIds') or [],
            },
            status='COMPLETED',
        )

        log_audit(
            request, module='EXPORT', action='EXPORT_PERFORMED',
            record_type='ExportHistory', record_id=history.id,
            summary=f"Exported {total} lead(s) to '{file_name}'.",
            new_values={'file_name': file_name, 'file_type': file_type, 'total_records': total},
        )

        return Response({
            'history': ExportHistorySerializer(history).data,
            'downloadUrl': f"/api/export-history/{history.id}/download/",
        }, status=status.HTTP_201_CREATED)


class ExportHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAdminUserRole]
    serializer_class = ExportHistorySerializer

    def get_queryset(self):
        qs = ExportHistory.objects.select_related('exported_by').order_by('-created_at')
        search = self.request.query_params.get('search')
        file_type = self.request.query_params.get('fileType')
        status_val = self.request.query_params.get('status')
        if search:
            qs = qs.filter(file_name__icontains=search)
        if file_type:
            qs = qs.filter(file_type=file_type.lower())
        if status_val:
            qs = qs.filter(status=status_val.upper())
        return qs

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        history = self.get_object()
        payload = history.filters_applied or {}
        filters = payload.get('filters') or {}
        qs = build_export_queryset(payload)
        leads = list(qs.order_by('-created_at')[:20000])
        rows = rows_from_leads(leads)

        try:
            return file_response(history.file_name, history.file_type, rows, filters)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:  # pragma: no cover
            history.status = 'FAILED'
            history.save(update_fields=['status'])
            return Response({'error': f'Failed to generate export: {exc}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExportStatsView(APIView):
    permission_classes = [IsAdminUserRole]

    def get(self, request):
        qs = ExportHistory.objects.all()
        total_exports = qs.count()
        today = timezone.now().date()
        exports_today = qs.filter(created_at__date=today).count()

        format_counts = qs.values('file_type').annotate(count=Count('id')).order_by('-count')
        most_exported_report = format_counts[0]['file_type'].upper() if format_counts else 'N/A'

        last = qs.order_by('-created_at').first()
        last_export_activity = last.created_at.isoformat() if last else None

        return Response({
            'totalExports': total_exports,
            'exportsToday': exports_today,
            'mostExportedReport': most_exported_report,
            'lastExportActivity': last_export_activity,
        })
